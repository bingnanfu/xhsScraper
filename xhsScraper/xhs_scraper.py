"""
小红书评论爬取工具

用法:
    python xhs_scraper.py login            # 首次使用，登录并保存 Cookie
    python xhs_scraper.py search <关键词>   # 搜索关键词并爬取前10篇笔记的评论

依赖:
    pip install playwright openpyxl
    playwright install chromium
"""

import asyncio
import json
import re
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path

from playwright.async_api import async_playwright, Page, Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ─── 配置 ──────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
COOKIE_FILE = SCRIPT_DIR / "cookies.json"
OUTPUT_DIR = SCRIPT_DIR / "output"

MAX_POSTS = 2               # 最多爬取笔记数
SCROLL_PAUSE = 1.5           # 每次滚动后等待秒数
STALE_LIMIT = 12             # 连续多少次滚动没有新评论后停止
POST_INTERVAL = 3            # 笔记间爬取间隔秒数

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)


# ─── Cookie 管理 ───────────────────────────────────────────────────────────────

async def cmd_login():
    """打开浏览器让用户手动登录，保存 Cookie"""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent=USER_AGENT,
        )
        page = await ctx.new_page()
        await page.goto("https://www.xiaohongshu.com")

        print()
        print("=" * 50)
        print("  请在浏览器中登录小红书账号")
        print("  登录成功后，回到此终端按 Enter 键继续")
        print("=" * 50)
        await asyncio.get_event_loop().run_in_executor(None, input)

        cookies = await ctx.cookies()
        COOKIE_FILE.write_text(
            json.dumps(cookies, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(f"Cookie 已保存（共 {len(cookies)} 条）")
        await browser.close()


def load_cookies() -> list[dict]:
    if not COOKIE_FILE.exists():
        print("错误: 未找到 Cookie 文件，请先运行: python xhs_scraper.py login")
        sys.exit(1)
    return json.loads(COOKIE_FILE.read_text(encoding="utf-8"))


# ─── 工具函数 ──────────────────────────────────────────────────────────────────

def parse_timestamp(ts) -> str:
    if not ts:
        return ""
    try:
        if isinstance(ts, str):
            return ts
        if isinstance(ts, (int, float)):
            if ts > 1e12:
                ts = ts / 1000
            return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        pass
    return str(ts)


# ─── 获取搜索页卡片 ───────────────────────────────────────────────────────────

async def _get_card_elements(page: Page):
    """获取搜索结果中的笔记卡片元素列表"""
    for selector in [
        "section.note-item",
        "[class*='note-item']",
        "div[class*='note-item']",
    ]:
        cards = await page.query_selector_all(selector)
        if cards:
            return cards
    return []


async def load_search_page(page: Page, keyword: str) -> int:
    """导航到搜索页，滚动加载卡片，返回卡片数量"""
    encoded = urllib.parse.quote(keyword)
    search_url = (
        f"https://www.xiaohongshu.com/search_result"
        f"?keyword={encoded}&source=web_search_result_notes"
    )
    await page.goto(search_url)
    await page.wait_for_load_state("networkidle")
    await asyncio.sleep(3)

    # 滚动加载更多卡片
    for _ in range(8):
        cards = await _get_card_elements(page)
        if len(cards) >= MAX_POSTS + 5:  # 多加载一些，预留去重空间
            break
        await page.evaluate("window.scrollBy(0, 1000)")
        await asyncio.sleep(2)

    # 回到顶部
    await page.evaluate("window.scrollTo(0, 0)")
    await asyncio.sleep(0.5)

    cards = await _get_card_elements(page)
    return len(cards)


# ─── 点击第 N 个卡片，爬取评论 ─────────────────────────────────────────────────

async def scrape_by_card_index(
    page: Page, card_index: int, visited_urls: set
) -> tuple[str, str, list[dict]]:
    """
    点击第 card_index 个卡片打开弹窗，爬取评论，关闭弹窗。
    如果打开的页面 URL 已在 visited_urls 中则跳过（去重）。
    返回 (标题, url, 评论列表)，重复页返回 url='DUPLICATE'。
    """
    comments = []
    seen_ids = set()
    post_title = ""
    post_url = ""
    main_has_more = True  # 主评论 API 是否还有更多页

    # ── 评论 API 拦截器 ──────────────────────────────────────
    async def on_comment_response(resp: Response):
        nonlocal main_has_more
        url = resp.url
        is_main = "/api/sns/web/v2/comment/page" in url and "/sub/" not in url
        is_sub = "/api/sns/web/v2/comment/sub/page" in url
        if not (is_main or is_sub):
            return
        try:
            body = await resp.json()
            data = body.get("data", {})
            if is_main and not data.get("has_more", True):
                main_has_more = False
            for c in data.get("comments", []):
                cid = c.get("id", "")
                if not cid or cid in seen_ids:
                    continue
                seen_ids.add(cid)
                user = c.get("user_info", {})
                comments.append({
                    "comment_id": cid,
                    "is_reply": is_sub,
                    "parent_comment_id": c.get("target_comment_id", ""),
                    "user_name": user.get("nickname", ""),
                    "content": c.get("content", ""),
                    "time": parse_timestamp(c.get("create_time")),
                    "like_count": int(c.get("like_count", 0) or 0),
                })
                for sub in c.get("sub_comments", []):
                    sub_id = sub.get("id", "")
                    if not sub_id or sub_id in seen_ids:
                        continue
                    seen_ids.add(sub_id)
                    sub_user = sub.get("user_info", {})
                    comments.append({
                        "comment_id": sub_id,
                        "is_reply": True,
                        "parent_comment_id": cid,
                        "user_name": sub_user.get("nickname", ""),
                        "content": sub.get("content", ""),
                        "time": parse_timestamp(sub.get("create_time")),
                        "like_count": int(sub.get("like_count", 0) or 0),
                    })
        except Exception as e:
            print(f"    [警告] 解析评论 API: {e}")

    page.on("response", on_comment_response)
    saved_search_url = page.url

    try:
        # ── 0) 确保在搜索页且无残留弹窗 ──────────────────────
        if "/explore/" in page.url:
            # 上一次的弹窗没关干净，强制回到搜索页
            await _close_note_detail(page, saved_search_url)
            await asyncio.sleep(1)

        await page.evaluate("""
        () => {
            document.querySelectorAll('.note-detail-mask').forEach(el => el.remove());
        }
        """)
        await asyncio.sleep(0.3)

        # 记录点击前的 URL，用于判断是否成功打开了新页面
        url_before_click = page.url

        # ── 1) 点击第 card_index 个卡片 ──────────────────────
        cards = await _get_card_elements(page)
        if card_index >= len(cards):
            return "", "", []

        card = cards[card_index]
        await card.scroll_into_view_if_needed()
        await asyncio.sleep(0.5)
        await card.click()
        await asyncio.sleep(2)

        # ── 2) 等待弹窗打开（URL 从搜索页变为 /explore/）────
        try:
            # 等待 URL 变化为一个新的 /explore/ 地址
            await page.wait_for_function(
                """(oldUrl) => {
                    const cur = window.location.href;
                    return cur.includes('/explore/') && cur !== oldUrl;
                }""",
                arg=url_before_click,
                timeout=8000,
            )
        except Exception:
            print("    [警告] 等待笔记详情超时，尝试继续...")

        await page.wait_for_load_state("networkidle")
        await asyncio.sleep(2)

        # ── 3) 去重检查：URL 是否已经爬过 ────────────────────
        post_url = page.url
        # 提取 /explore/{id} 部分作为去重 key
        match = re.search(r"/explore/([a-f0-9]+)", post_url)
        dedup_key = match.group(1) if match else post_url

        if dedup_key in visited_urls:
            print(f"    [跳过] 该笔记已爬取过，关闭弹窗...")
            await _close_note_detail(page, saved_search_url)
            return "", "DUPLICATE", []

        visited_urls.add(dedup_key)

        # ── 4) 获取标题 ──────────────────────────────────────
        post_title = await _get_note_title(page)

        # ── 4.5) 读取评论总数（用于进度参考）─────────────────
        total_text = await page.evaluate("""
        () => {
            const el = document.querySelector('.comments-container .total');
            return el ? el.innerText : '';
        }
        """)
        total_match = re.search(r"(\d+)", total_text or "")
        total_comments = int(total_match.group(1)) if total_match else 0
        if total_comments > 0:
            print(f"    页面显示共 {total_comments} 条评论")

        # ── 5) 智能滚动加载评论（按增量驱动）────────────────
        prev_count = 0
        stale_rounds = 0
        scroll_round = 0

        while stale_rounds < STALE_LIMIT:
            # 主评论 API 已明确返回 has_more=false，停止滚动
            if not main_has_more:
                print(f"    主评论 API 已无更多数据，停止滚动")
                break

            await _do_scroll(page)
            await asyncio.sleep(SCROLL_PAUSE)
            await _click_more_comments(page)
            scroll_round += 1

            current = len(comments)
            if current == prev_count:
                stale_rounds += 1
            else:
                stale_rounds = 0
                prev_count = current

            if scroll_round % 10 == 0:
                main_count = sum(1 for c in comments if not c.get("is_reply"))
                progress = f"/{total_comments}" if total_comments else ""
                print(f"    ... 已滚动 {scroll_round} 轮, 当前 {current} 条 (主评论 {main_count}{progress})")

        print(f"    评论滚动完毕 ({scroll_round} 轮)")

        # ── 6) 多轮展开子回复 ────────────────────────────────
        expand_round = 0
        while expand_round < 20:
            prev = len(comments)
            clicked = await _click_reply_buttons(page)
            if not clicked:
                break
            await asyncio.sleep(2)
            await _do_scroll(page)
            await asyncio.sleep(1)
            expand_round += 1
            new_count = len(comments) - prev
            if new_count > 0:
                print(f"    ... 展开回复第 {expand_round} 轮, 新增 {new_count} 条")

        # 等待最后的 API 响应
        await asyncio.sleep(2)

        # ── 7) 关闭弹窗 ─────────────────────────────────────
        await _close_note_detail(page, saved_search_url)

    except Exception as e:
        print(f"    [错误] 爬取异常: {e}")
        try:
            await _close_note_detail(page, saved_search_url)
        except Exception:
            pass
    finally:
        page.remove_listener("response", on_comment_response)

    # 给每条评论补上 post 信息
    for c in comments:
        c["post_title"] = post_title
        c["post_url"] = post_url

    return post_title, post_url, comments


# ─── 滚动 & 按钮操作 ──────────────────────────────────────────────────────────

async def _do_scroll(page: Page):
    """滚动 .note-scroller 容器到底部以触发懒加载"""
    await page.evaluate("""
    () => {
        const scroller = document.querySelector('.note-scroller');
        if (scroller) {
            scroller.scrollTop = scroller.scrollHeight;
            return;
        }
        window.scrollTo(0, document.body.scrollHeight);
    }
    """)


async def _get_note_title(page: Page) -> str:
    for selector in [
        "#detail-title",
        "[class*='title'][class*='note']",
        "[class*='note-title']",
        "[class*='detail'] [class*='title']",
        ".title",
    ]:
        try:
            el = await page.query_selector(selector)
            if el:
                text = (await el.inner_text()).strip()
                if text and len(text) > 1:
                    return text
        except Exception:
            pass
    try:
        title = await page.title()
        if title:
            title = re.sub(r'\s*[-|–]\s*小红书.*$', '', title).strip()
            if title:
                return title
    except Exception:
        pass
    return "无标题"


async def _click_more_comments(page: Page):
    for pattern in ["查看更多评论", "展开更多评论", "更多评论", "点击加载"]:
        try:
            btns = await page.query_selector_all(f"text=/{pattern}/")
            for btn in btns:
                try:
                    if await btn.is_visible():
                        await btn.click()
                        await asyncio.sleep(1)
                except Exception:
                    pass
        except Exception:
            pass
    for sel in ["[class*='show-more']", "[class*='load-more']"]:
        try:
            btns = await page.query_selector_all(sel)
            for btn in btns:
                try:
                    if await btn.is_visible():
                        await btn.click()
                        await asyncio.sleep(1)
                except Exception:
                    pass
        except Exception:
            pass


async def _click_reply_buttons(page: Page) -> bool:
    clicked = False
    for pattern in [
        "展开.*回复",
        "查看.*回复",
        "共.*条回复",
        "展开更多",
    ]:
        try:
            btns = await page.query_selector_all(f"text=/{pattern}/")
            for btn in btns:
                try:
                    if await btn.is_visible():
                        await btn.click()
                        await asyncio.sleep(1.5)
                        clicked = True
                except Exception:
                    pass
        except Exception:
            pass
    for sel in [
        "[class*='reply'] [class*='more']",
        "[class*='show-more']",
        "[class*='expand']",
    ]:
        try:
            elements = await page.query_selector_all(sel)
            for el in elements:
                try:
                    if await el.is_visible():
                        text = await el.inner_text()
                        if any(kw in text for kw in ["展开", "查看", "回复", "更多"]):
                            await el.click()
                            await asyncio.sleep(1.5)
                            clicked = True
                except Exception:
                    pass
        except Exception:
            pass
    return clicked


async def _close_note_detail(page: Page, search_url: str):
    """关闭笔记详情弹窗（.note-detail-mask），回到搜索结果页"""
    if "search_result" in page.url and "/explore/" not in page.url:
        return

    methods = [
        ("关闭按钮", _try_close_btn),
        ("Escape", _try_escape),
        ("浏览器后退", _try_go_back),
        ("JS移除+后退", _try_js_remove_and_back),
        ("直接导航", _try_navigate_back),
    ]

    for name, method in methods:
        try:
            await method(page, search_url)
            await asyncio.sleep(1)
            if "search_result" in page.url and "/explore/" not in page.url:
                break
        except Exception:
            pass

    # 最终保障：确保弹窗已移除（防止残留遮挡下次点击）
    await page.evaluate("""
    () => {
        document.querySelectorAll('.note-detail-mask').forEach(el => el.remove());
    }
    """)
    await asyncio.sleep(0.5)

    # 最终检查：如果仍不在搜索页，强制导航
    if "search_result" not in page.url or "/explore/" in page.url:
        if search_url:
            await page.goto(search_url)
            await page.wait_for_load_state("networkidle")
            await asyncio.sleep(2)


async def _try_close_btn(page: Page, search_url: str):
    close_btn = await page.query_selector(".note-detail-mask [class*='close']")
    if close_btn and await close_btn.is_visible():
        await close_btn.click()


async def _try_escape(page: Page, search_url: str):
    await page.keyboard.press("Escape")


async def _try_go_back(page: Page, search_url: str):
    await page.go_back()


async def _try_js_remove_and_back(page: Page, search_url: str):
    await page.evaluate("""
    () => {
        const mask = document.querySelector('.note-detail-mask');
        if (mask) mask.remove();
    }
    """)
    await asyncio.sleep(0.5)
    if "/explore/" in page.url:
        await page.go_back()


async def _try_navigate_back(page: Page, search_url: str):
    if search_url:
        await page.goto(search_url)
        await page.wait_for_load_state("networkidle")


# ─── 导出 Excel ────────────────────────────────────────────────────────────────

def export_excel(comments: list[dict], keyword: str) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = OUTPUT_DIR / f"xhs_{keyword}_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "评论数据"

    headers = [
        "笔记标题", "笔记链接", "评论类型", "用户名",
        "评论内容", "时间", "点赞数", "父评论ID",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, c in enumerate(comments, 2):
        values = [
            c.get("post_title", ""),
            c.get("post_url", ""),
            "回复" if c.get("is_reply") else "评论",
            c.get("user_name", ""),
            c.get("content", ""),
            c.get("time", ""),
            c.get("like_count", 0),
            c.get("parent_comment_id", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="top", wrap_text=(col_idx == 5),
            )
            if col_idx == 7:
                cell.alignment = Alignment(horizontal="center")

    col_widths = {"A": 30, "B": 45, "C": 10, "D": 15,
                  "E": 60, "F": 20, "G": 10, "H": 22}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(comments) + 1}"
    wb.save(filepath)
    return filepath


# ─── 主流程 ────────────────────────────────────────────────────────────────────

async def cmd_search(keyword: str):
    cookies = load_cookies()
    OUTPUT_DIR.mkdir(exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=200)
        ctx = await browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent=USER_AGENT,
        )
        await ctx.add_cookies(cookies)
        page = await ctx.new_page()

        # 1) 搜索并加载卡片
        print(f"\n正在搜索「{keyword}」...")
        card_count = await load_search_page(page, keyword)
        if card_count == 0:
            print("未找到任何笔记，请检查关键词或重新登录")
            await browser.close()
            return

        print(f"搜索页加载了 {card_count} 个卡片")

        # 2) 逐个点击卡片，遇到重复的跳过，直到爬够 MAX_POSTS 篇
        all_comments = []
        visited_urls = set()    # 已爬取的笔记 ID，用于去重
        scraped_count = 0       # 成功爬取的笔记数
        card_index = 0          # 当前要点击的卡片序号

        while scraped_count < MAX_POSTS and card_index < card_count + 10:
            print(f"\n[已完成 {scraped_count}/{MAX_POSTS}] 点击第 {card_index + 1} 个卡片...")

            try:
                title, url, comments = await scrape_by_card_index(
                    page, card_index, visited_urls
                )

                if url == "DUPLICATE":
                    print(f"    重复，跳过")
                    card_index += 1
                    continue

                if url == "":
                    # 没有更多卡片了
                    print(f"    没有更多卡片")
                    break

                if comments:
                    all_comments.extend(comments)
                    ccount = sum(1 for c in comments if not c.get("is_reply"))
                    rcount = sum(1 for c in comments if c.get("is_reply"))
                    print(f"  标题: {title[:50]}")
                    print(f"  获取 {ccount} 条评论 + {rcount} 条回复")
                else:
                    print(f"  标题: {title[:50] if title else '无标题'}")
                    print(f"  未获取到评论")

                scraped_count += 1

            except Exception as e:
                print(f"  [错误] 爬取失败: {e}")

            card_index += 1

            if scraped_count < MAX_POSTS:
                print(f"  等待 {POST_INTERVAL} 秒...")
                await asyncio.sleep(POST_INTERVAL)

        # 3) 导出
        if all_comments:
            filepath = export_excel(all_comments, keyword)
            print()
            print("=" * 50)
            print(f"  爬取完成!")
            print(f"  共爬取 {scraped_count} 篇笔记")
            print(f"  共获取 {len(all_comments)} 条评论/回复")
            print(f"  文件: {filepath}")
            print("=" * 50)
        else:
            print("\n未获取到任何评论数据，可能原因:")
            print("  1. Cookie 已过期，请重新 login")
            print("  2. 小红书页面结构有变化")
            print("  3. 被反爬机制拦截")

        await browser.close()


def main():
    if len(sys.argv) < 2:
        print("小红书评论爬取工具")
        print()
        print("用法:")
        print("  python xhs_scraper.py login            登录并保存 Cookie")
        print("  python xhs_scraper.py search <关键词>   搜索并爬取评论")
        print()
        print("示例:")
        print("  python xhs_scraper.py login")
        print('  python xhs_scraper.py search "咖啡推荐"')
        sys.exit(0)

    cmd = sys.argv[1].lower()
    if cmd == "login":
        asyncio.run(cmd_login())
    elif cmd == "search":
        if len(sys.argv) < 3:
            kw = input("请输入搜索关键词: ").strip()
        else:
            kw = " ".join(sys.argv[2:])
        if not kw:
            print("关键词不能为空")
            sys.exit(1)
        asyncio.run(cmd_search(kw))
    else:
        print(f"未知命令: {cmd}")
        print("可用命令: login, search")
        sys.exit(1)


if __name__ == "__main__":
    main()
